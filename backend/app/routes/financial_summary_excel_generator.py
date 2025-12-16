from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import tempfile
from typing import Dict, Any, List

class FinancialSummaryExcelGenerator:
    """Handles Excel report generation for Financial Summary views"""
    
    # Shared styles
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TOTAL_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    SUBTITLE_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    @staticmethod
    def format_currency(amount: float) -> float:
        """Return full amount (not in Lakhs)"""
        if amount is None:
            return 0.0
        return round(amount, 2)
    
    @staticmethod
    def format_date(date_value) -> str:
        """Format date to DD-MMM-YYYY"""
        if not date_value:
            return ""
        try:
            if isinstance(date_value, str):
                dt = datetime.fromisoformat(date_value.replace('Z', '+00:00'))
            else:
                dt = date_value
            return dt.strftime("%d-%b-%Y")
        except:
            return str(date_value)
    
    @staticmethod
    def generate_financial_summary_excel(
        view_mode: str,
        data: List[Dict[str, Any]],
        summary: Dict[str, Any],
        filters: Dict[str, Any]
    ) -> str:
        """Generate Excel report based on view mode"""
        
        wb = Workbook()
        ws = wb.active
        
        # Generate based on view mode
        if view_mode == 'by_project':
            FinancialSummaryExcelGenerator._create_by_project_sheet(ws, data, summary, filters)
        elif view_mode == 'by_budget_head':
            FinancialSummaryExcelGenerator._create_by_budget_head_sheet(ws, data, summary, filters)
        elif view_mode == 'by_technical_group':
            FinancialSummaryExcelGenerator._create_by_technical_group_sheet(ws, data, summary, filters)
        elif view_mode == 'by_funding_agency':
            FinancialSummaryExcelGenerator._create_by_funding_agency_sheet(ws, data, summary, filters)
        elif view_mode == 'project_budget_head_detail':
            FinancialSummaryExcelGenerator._create_project_budget_head_detail_sheet(ws, data, summary, filters)
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        temp_file.close()
        
        wb.save(temp_path)
        return temp_path
    
    @staticmethod
    def _add_header_and_filters(ws, title: str, filters: Dict[str, Any]):
        """Add common header and filter information"""
        # Title
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True, color="1F2937")
        ws.merge_cells('A1:J1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Generated timestamp
        row = 2
        ws[f'A{row}'] = f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
        ws[f'A{row}'].font = Font(size=10, italic=True, color="6B7280")
        ws.merge_cells(f'A{row}:J{row}')
        
        # Filter information
        row = 3
        filter_text = FinancialSummaryExcelGenerator._get_filter_text(filters)
        if filter_text:
            ws[f'A{row}'] = f"Filters: {filter_text}"
            ws[f'A{row}'].font = Font(size=10, color="6B7280")
            ws.merge_cells(f'A{row}:J{row}')
            row += 1
        
        return row + 1  # Return next available row
    
    @staticmethod
    def _get_filter_text(filters: Dict[str, Any]) -> str:
        """Generate filter description text"""
        parts = []
        
        date_mode = filters.get('dateFilterMode', 'current')
        if date_mode == 'as_of_date' and filters.get('asOfDate'):
            parts.append(f"As of {filters['asOfDate']}")
        elif date_mode == 'date_range' and filters.get('startDate') and filters.get('endDate'):
            parts.append(f"From {filters['startDate']} to {filters['endDate']}")
        elif date_mode == 'financial_year' and filters.get('financialYear'):
            parts.append(f"FY {filters['financialYear']}-{int(filters['financialYear'])+1}")
        elif date_mode == 'monthly' and filters.get('year') and filters.get('month'):
            parts.append(f"{datetime(int(filters['year']), int(filters['month']), 1).strftime('%B %Y')}")
        elif date_mode == 'quarterly' and filters.get('year') and filters.get('quarter'):
            parts.append(f"Q{filters['quarter']} {filters['year']}")
        else:
            parts.append("All Time")
        
        if filters.get('projectName'):
            parts.append(f"Project: {filters['projectName']}")
        
        return " | ".join(parts)
    
    @staticmethod
    def _create_by_project_sheet(ws, data: List[Dict], summary: Dict, filters: Dict):
        """Create By Project view sheet"""
        ws.title = "By Project"
        
        # Add header
        start_row = FinancialSummaryExcelGenerator._add_header_and_filters(
            ws, "FINANCIAL SUMMARY - BY PROJECT", filters
        )
        
        # Summary section
        ws[f'A{start_row}'] = "OVERALL SUMMARY"
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        ws[f'A{start_row}'].fill = FinancialSummaryExcelGenerator.SUBTITLE_FILL
        ws.merge_cells(f'A{start_row}:D{start_row}')
        
        start_row += 1
        summary_items = [
            ('Total Projects:', summary.get('total_projects', 0)),
            ('Total Approved Budget:', FinancialSummaryExcelGenerator.format_currency(summary.get('total_approved_budget', 0))),
            ('Total Funds Received:', FinancialSummaryExcelGenerator.format_currency(summary.get('total_funds_received', 0))),
            ('Total Expenditure:', FinancialSummaryExcelGenerator.format_currency(summary.get('total_expenditure', 0))),
            ('Budget Balance:', FinancialSummaryExcelGenerator.format_currency(summary.get('budget_balance', 0))),
            ('Funds Balance:', FinancialSummaryExcelGenerator.format_currency(summary.get('funds_balance', 0))),
        ]
        
        for label, value in summary_items:
            ws[f'A{start_row}'] = label
            ws[f'A{start_row}'].font = Font(bold=True)
            ws[f'B{start_row}'] = value
            if isinstance(value, (int, float)) and label != 'Total Projects:':
                ws[f'B{start_row}'].number_format = '₹#,##0.00'
            start_row += 1
        
        # Table header
        start_row += 2
        headers = [
            'Project No', 'Title', 'Technical Group', 'Funding Agency',
            'Approved Budget', 'Funds Received', 'Expenditure',
            'Budget Balance', 'Funds Balance', 'Utilization %'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = FinancialSummaryExcelGenerator.HEADER_FONT
            cell.fill = FinancialSummaryExcelGenerator.HEADER_FILL
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Data rows
        row = start_row + 1
        totals = {
            'approved_budget': 0,
            'funds_received': 0,
            'expenditure': 0,
            'budget_balance': 0,
            'funds_balance': 0
        }
        
        for project in data:
            # Handle both direct values and computed values
            approved = float(project.get('approved_budget', 0))
            funds = float(project.get('funds_received', 0))
            expenditure = float(project.get('expenditure', 0))
            budget_balance = float(project.get('budget_balance', 0))
            funds_balance = float(project.get('funds_balance', 0))
            utilization = float(project.get('utilization_percentage', 0))
            
            totals['approved_budget'] += approved
            totals['funds_received'] += funds
            totals['expenditure'] += expenditure
            totals['budget_balance'] += budget_balance
            totals['funds_balance'] += funds_balance
            
            ws[f'A{row}'] = project.get('project_no', 'N/A')
            ws[f'B{row}'] = project.get('title', 'N/A')
            ws[f'C{row}'] = project.get('technical_group', 'N/A')
            ws[f'D{row}'] = project.get('funding_agency', 'N/A')
            ws[f'E{row}'] = approved
            ws[f'F{row}'] = funds
            ws[f'G{row}'] = expenditure
            ws[f'H{row}'] = budget_balance
            ws[f'I{row}'] = funds_balance
            ws[f'J{row}'] = utilization / 100  # Convert to decimal for percentage format
            
            # Formatting
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
                cell.alignment = Alignment(vertical='center')
            
            # Currency formatting
            for col in ['E', 'F', 'G', 'H', 'I']:
                ws[f'{col}{row}'].number_format = '₹#,##0.00'
            
            ws[f'J{row}'].number_format = '0.0%'
            
            # Color coding for balances
            if budget_balance > 0:
                ws[f'H{row}'].font = Font(color="10B981")
            elif budget_balance < 0:
                ws[f'H{row}'].font = Font(color="EF4444")
            
            if funds_balance > 0:
                ws[f'I{row}'].font = Font(color="10B981")
            elif funds_balance < 0:
                ws[f'I{row}'].font = Font(color="EF4444")
            
            row += 1
        
        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'E{row}'] = totals['approved_budget']
        ws[f'F{row}'] = totals['funds_received']
        ws[f'G{row}'] = totals['expenditure']
        ws[f'H{row}'] = totals['budget_balance']
        ws[f'I{row}'] = totals['funds_balance']
        
        avg_utilization = (totals['expenditure'] / totals['funds_received']) if totals['funds_received'] > 0 else 0
        ws[f'J{row}'] = avg_utilization
        
        # Style total row
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = FinancialSummaryExcelGenerator.TOTAL_FILL
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
        
        for col in ['E', 'F', 'G', 'H', 'I']:
            ws[f'{col}{row}'].number_format = '₹#,##0.00'
        ws[f'J{row}'].number_format = '0.0%'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 18
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 14
        
        # Freeze panes
        ws.freeze_panes = f'A{start_row + 1}'
    
    @staticmethod
    def _create_by_budget_head_sheet(ws, data: List[Dict], summary: Dict, filters: Dict):
        """Create By Budget Head view sheet"""
        ws.title = "By Budget Head"
        
        start_row = FinancialSummaryExcelGenerator._add_header_and_filters(
            ws, "FINANCIAL SUMMARY - BY BUDGET HEAD", filters
        )
        
        # Table header
        headers = [
            'Budget Head', 'Approved Budget', 'Funds Received', 'Expenditure',
            'Budget Balance', 'Funds Balance', 'Utilization %'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = FinancialSummaryExcelGenerator.HEADER_FONT
            cell.fill = FinancialSummaryExcelGenerator.HEADER_FILL
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Budget head order
        budget_head_order = [
            'manpower', 'equipment', 'travel & training',
            'consumables', 'contingency', 'overhead'
        ]
        
        # Sort data by budget head order
        sorted_data = sorted(data, key=lambda x: 
            budget_head_order.index(x.get('budget_head', '').lower()) 
            if x.get('budget_head', '').lower() in budget_head_order else 999
        )
        
        # Data rows
        row = start_row + 1
        totals = {
            'approved_budget': 0,
            'funds_received': 0,
            'expenditure': 0,
            'budget_balance': 0,
            'funds_balance': 0
        }
        
        for item in sorted_data:
            # Handle both field naming conventions
            approved = float(item.get('approved_budget') or item.get('total_approved', 0))
            funds = float(item.get('funds_received') or item.get('total_funds_received', 0))
            expenditure = float(item.get('expenditure') or item.get('total_expenditure', 0))
            budget_balance = float(item.get('budget_balance', 0))
            funds_balance = float(item.get('funds_balance', 0))
            utilization = float(item.get('utilization_percentage', 0))
            
            totals['approved_budget'] += approved
            totals['funds_received'] += funds
            totals['expenditure'] += expenditure
            totals['budget_balance'] += budget_balance
            totals['funds_balance'] += funds_balance
            
            ws[f'A{row}'] = str(item.get('budget_head', 'N/A')).title()
            ws[f'B{row}'] = approved
            ws[f'C{row}'] = funds
            ws[f'D{row}'] = expenditure
            ws[f'E{row}'] = budget_balance
            ws[f'F{row}'] = funds_balance
            ws[f'G{row}'] = utilization / 100
            
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
            
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row}'].number_format = '₹#,##0.00'
            ws[f'G{row}'].number_format = '0.0%'
            
            row += 1
        
        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = totals['approved_budget']
        ws[f'C{row}'] = totals['funds_received']
        ws[f'D{row}'] = totals['expenditure']
        ws[f'E{row}'] = totals['budget_balance']
        ws[f'F{row}'] = totals['funds_balance']
        avg_util = (totals['expenditure'] / totals['funds_received']) if totals['funds_received'] > 0 else 0
        ws[f'G{row}'] = avg_util
        
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = FinancialSummaryExcelGenerator.TOTAL_FILL
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
        
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].number_format = '₹#,##0.00'
        ws[f'G{row}'].number_format = '0.0%'
        
        # Column widths
        ws.column_dimensions['A'].width = 22
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['G'].width = 16
    
    @staticmethod
    def _create_by_technical_group_sheet(ws, data: List[Dict], summary: Dict, filters: Dict):
        """Create By Technical Group view sheet"""
        ws.title = "By Technical Group"
        
        start_row = FinancialSummaryExcelGenerator._add_header_and_filters(
            ws, "FINANCIAL SUMMARY - BY TECHNICAL GROUP", filters
        )
        
        headers = [
            'Technical Group', 'Projects', 'Approved Budget', 'Funds Received',
            'Expenditure', 'Budget Balance', 'Funds Balance', 'Utilization %'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = FinancialSummaryExcelGenerator.HEADER_FONT
            cell.fill = FinancialSummaryExcelGenerator.HEADER_FILL
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        row = start_row + 1
        totals = {
            'projects': 0,
            'approved_budget': 0,
            'funds_received': 0,
            'expenditure': 0,
            'budget_balance': 0,
            'funds_balance': 0
        }
        
        for item in data:
            # Handle both field naming conventions
            approved = float(item.get('approved_budget') or item.get('total_approved', 0))
            funds = float(item.get('funds_received') or item.get('total_funds_received', 0))
            expenditure = float(item.get('expenditure') or item.get('total_expenditure', 0))
            budget_balance = float(item.get('budget_balance', 0))
            funds_balance = float(item.get('funds_balance', 0))
            utilization = float(item.get('utilization_percentage', 0))
            project_count = int(item.get('project_count', 0))
            
            totals['projects'] += project_count
            totals['approved_budget'] += approved
            totals['funds_received'] += funds
            totals['expenditure'] += expenditure
            totals['budget_balance'] += budget_balance
            totals['funds_balance'] += funds_balance
            
            # Handle group_name field
            group_name = item.get('technical_group') or item.get('group_name', 'N/A')
            
            ws[f'A{row}'] = group_name
            ws[f'B{row}'] = project_count
            ws[f'C{row}'] = approved
            ws[f'D{row}'] = funds
            ws[f'E{row}'] = expenditure
            ws[f'F{row}'] = budget_balance
            ws[f'G{row}'] = funds_balance
            ws[f'H{row}'] = utilization / 100
            
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
            
            for col in ['C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].number_format = '₹#,##0.00'
            ws[f'H{row}'].number_format = '0.0%'
            
            row += 1
        
        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = totals['projects']
        ws[f'C{row}'] = totals['approved_budget']
        ws[f'D{row}'] = totals['funds_received']
        ws[f'E{row}'] = totals['expenditure']
        ws[f'F{row}'] = totals['budget_balance']
        ws[f'G{row}'] = totals['funds_balance']
        avg_util = (totals['expenditure'] / totals['funds_received']) if totals['funds_received'] > 0 else 0
        ws[f'H{row}'] = avg_util
        
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = FinancialSummaryExcelGenerator.TOTAL_FILL
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
        
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row}'].number_format = '₹#,##0.00'
        ws[f'H{row}'].number_format = '0.0%'
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['H'].width = 16
    
    @staticmethod
    def _create_by_funding_agency_sheet(ws, data: List[Dict], summary: Dict, filters: Dict):
        """Create By Funding Agency view sheet"""
        ws.title = "By Funding Agency"
        
        start_row = FinancialSummaryExcelGenerator._add_header_and_filters(
            ws, "FINANCIAL SUMMARY - BY FUNDING AGENCY", filters
        )
        
        headers = [
            'Funding Agency', 'Projects', 'Approved Budget', 'Funds Received',
            'Expenditure', 'Budget Balance', 'Funds Balance', 'Utilization %'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = FinancialSummaryExcelGenerator.HEADER_FONT
            cell.fill = FinancialSummaryExcelGenerator.HEADER_FILL
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        row = start_row + 1
        totals = {
            'projects': 0,
            'approved_budget': 0,
            'funds_received': 0,
            'expenditure': 0,
            'budget_balance': 0,
            'funds_balance': 0
        }
        
        for item in data:
            # Handle both field naming conventions
            approved = float(item.get('approved_budget') or item.get('total_approved', 0))
            funds = float(item.get('funds_received') or item.get('total_funds_received', 0))
            expenditure = float(item.get('expenditure') or item.get('total_expenditure', 0))
            budget_balance = float(item.get('budget_balance', 0))
            funds_balance = float(item.get('funds_balance', 0))
            utilization = float(item.get('utilization_percentage', 0))
            project_count = int(item.get('project_count', 0))
            
            totals['projects'] += project_count
            totals['approved_budget'] += approved
            totals['funds_received'] += funds
            totals['expenditure'] += expenditure
            totals['budget_balance'] += budget_balance
            totals['funds_balance'] += funds_balance
            
            # Handle agency_name field
            agency_name = item.get('funding_agency') or item.get('agency_name', 'N/A')
            
            ws[f'A{row}'] = agency_name
            ws[f'B{row}'] = project_count
            ws[f'C{row}'] = approved
            ws[f'D{row}'] = funds
            ws[f'E{row}'] = expenditure
            ws[f'F{row}'] = budget_balance
            ws[f'G{row}'] = funds_balance
            ws[f'H{row}'] = utilization / 100
            
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
            
            for col in ['C', 'D', 'E', 'F', 'G']:
                ws[f'{col}{row}'].number_format = '₹#,##0.00'
            ws[f'H{row}'].number_format = '0.0%'
            
            row += 1
        
        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = totals['projects']
        ws[f'C{row}'] = totals['approved_budget']
        ws[f'D{row}'] = totals['funds_received']
        ws[f'E{row}'] = totals['expenditure']
        ws[f'F{row}'] = totals['budget_balance']
        ws[f'G{row}'] = totals['funds_balance']
        avg_util = (totals['expenditure'] / totals['funds_received']) if totals['funds_received'] > 0 else 0
        ws[f'H{row}'] = avg_util
        
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = FinancialSummaryExcelGenerator.TOTAL_FILL
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
        
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{row}'].number_format = '₹#,##0.00'
        ws[f'H{row}'].number_format = '0.0%'
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        for col in ['C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['H'].width = 16
    
    @staticmethod
    def _create_project_budget_head_detail_sheet(ws, data: List[Dict], summary: Dict, filters: Dict):
        """Create Project Budget Head Detail view sheet"""
        ws.title = "Budget Head Detail"
        
        project_name = filters.get('projectName', 'Unknown Project')
        
        start_row = FinancialSummaryExcelGenerator._add_header_and_filters(
            ws, f"PROJECT BUDGET HEAD DETAIL - {project_name}", filters
        )
        
        headers = [
            'Budget Head', 'Approved Budget', 'Funds Received', 'Expenditure',
            'Budget Balance', 'Funds Balance', 'Utilization %'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = FinancialSummaryExcelGenerator.HEADER_FONT
            cell.fill = FinancialSummaryExcelGenerator.HEADER_FILL
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Budget head order
        budget_head_order = [
            'manpower', 'equipment', 'travel & training',
            'consumables', 'contingency', 'overhead'
        ]
        
        # Sort data
        sorted_data = sorted(data, key=lambda x: 
            budget_head_order.index(x.get('budget_head', '').lower()) 
            if x.get('budget_head', '').lower() in budget_head_order else 999
        )
        
        row = start_row + 1
        totals = {
            'approved_budget': 0,
            'funds_received': 0,
            'expenditure': 0,
            'budget_balance': 0,
            'funds_balance': 0
        }
        
        for item in sorted_data:
            approved = FinancialSummaryExcelGenerator.format_currency(item.get('approved_budget', 0))
            funds = FinancialSummaryExcelGenerator.format_currency(item.get('funds_received', 0))
            expenditure = FinancialSummaryExcelGenerator.format_currency(item.get('expenditure', 0))
            budget_balance = FinancialSummaryExcelGenerator.format_currency(item.get('budget_balance', 0))
            funds_balance = FinancialSummaryExcelGenerator.format_currency(item.get('funds_balance', 0))
            utilization = round(item.get('utilization_percentage', 0), 1)
            
            totals['approved_budget'] += approved
            totals['funds_received'] += funds
            totals['expenditure'] += expenditure
            totals['budget_balance'] += budget_balance
            totals['funds_balance'] += funds_balance
            
            ws[f'A{row}'] = str(item.get('budget_head', 'N/A')).title()
            ws[f'B{row}'] = approved
            ws[f'C{row}'] = funds
            ws[f'D{row}'] = expenditure
            ws[f'E{row}'] = budget_balance
            ws[f'F{row}'] = funds_balance
            ws[f'G{row}'] = utilization / 100
            
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
            
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row}'].number_format = '₹#,##0.00'
            ws[f'G{row}'].number_format = '0.0%'
            
            # Add breakdown if available
            if item.get('breakdown'):
                row += 1
                FinancialSummaryExcelGenerator._add_breakdown_section(ws, row, item['breakdown'])
                row += len(item['breakdown']) + 2  # Move past breakdown
            else:
                row += 1
        
        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = totals['approved_budget']
        ws[f'C{row}'] = totals['funds_received']
        ws[f'D{row}'] = totals['expenditure']
        ws[f'E{row}'] = totals['budget_balance']
        ws[f'F{row}'] = totals['funds_balance']
        avg_util = (totals['expenditure'] / totals['funds_received']) if totals['funds_received'] > 0 else 0
        ws[f'G{row}'] = avg_util
        
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = FinancialSummaryExcelGenerator.TOTAL_FILL
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
        
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].number_format = '₹#,##0.00'
        ws[f'G{row}'].number_format = '0.0%'
        
        # Column widths
        ws.column_dimensions['A'].width = 22
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
        ws.column_dimensions['G'].width = 16
    
    @staticmethod
    def _add_breakdown_section(ws, start_row: int, breakdown: List[Dict]):
        """Add item-wise breakdown section"""
        # Breakdown header
        ws.merge_cells(f'A{start_row}:G{start_row}')
        ws[f'A{start_row}'] = "Item-wise Breakdown:"
        ws[f'A{start_row}'].font = Font(size=10, bold=True, italic=True, color="4B5563")
        ws[f'A{start_row}'].fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        # Breakdown table headers
        start_row += 1
        breakdown_headers = ['Item', 'Approved Budget', 'Funds Received', 'Expenditure', 'Budget Balance', 'Funds Balance', 'Util %']
        
        for col, header in enumerate(breakdown_headers, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.value = header
            cell.font = Font(size=9, bold=True)
            cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
            cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # Breakdown data
        row = start_row + 1
        for item in breakdown:
            ws[f'A{row}'] = item.get('item_name', 'N/A')
            ws[f'B{row}'] = FinancialSummaryExcelGenerator.format_currency(item.get('approved_budget', 0))
            ws[f'C{row}'] = FinancialSummaryExcelGenerator.format_currency(item.get('funds_received', 0))
            ws[f'D{row}'] = FinancialSummaryExcelGenerator.format_currency(item.get('expenditure', 0))
            ws[f'E{row}'] = FinancialSummaryExcelGenerator.format_currency(item.get('budget_balance', 0))
            ws[f'F{row}'] = FinancialSummaryExcelGenerator.format_currency(item.get('funds_balance', 0))
            ws[f'G{row}'] = round(item.get('utilization_percentage', 0), 1) / 100
            
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = FinancialSummaryExcelGenerator.THIN_BORDER
                cell.font = Font(size=9)
            
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{row}'].number_format = '₹#,##0.00'
            ws[f'G{row}'].number_format = '0.0%'
            
            row += 1