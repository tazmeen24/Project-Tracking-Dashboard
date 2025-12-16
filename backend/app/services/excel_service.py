from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import tempfile
from typing import Dict, Any

class ExcelReportGenerator:
    """Handles Excel report generation"""
    
    @staticmethod
    def format_currency(amount: float) -> float:
        """Return full amount (not in Lakhs)"""
        if amount is None:
            return 0.0
        return round(amount, 2)
    
    @staticmethod
    def format_date(date_value) -> str:
        """Format date to DD-MMM-YYYY"""
        if not date_value:
            return ""
        try:
            if isinstance(date_value, str):
                dt = datetime.fromisoformat(date_value.replace('Z', '+00:00'))
            else:
                dt = date_value
            return dt.strftime("%d-%b-%Y")
        except:
            return str(date_value)
    
    @staticmethod
    def generate_excel(project_data: Dict[str, Any], include_sections: Dict[str, bool]) -> str:
        """Generate Excel report and return file path"""
        
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets based on included sections
        if include_sections.get('financial_summary', True):
            ExcelReportGenerator._create_summary_sheet(wb, project_data)
        
        if include_sections.get('budget_allocation', True):
            ExcelReportGenerator._create_budget_sheet(wb, project_data)
        
        if include_sections.get('funds_expenditure', True):
            ExcelReportGenerator._create_funds_sheet(wb, project_data)
        
        if include_sections.get('detailed_transactions', False):
            ExcelReportGenerator._create_transaction_sheets(wb, project_data)
        
        # Save to temporary file - let Python choose temp directory automatically
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        temp_file.close()
        
        wb.save(temp_path)
        return temp_path
    
    @staticmethod
    def _create_summary_sheet(wb: Workbook, data: Dict[str, Any]):
        """Create Summary sheet"""
        
        ws = wb.create_sheet("Summary", 0)
        project = data.get('project', {})
        summary = data.get('financial_summary', {})
        
        # Title
        ws['A1'] = "PROJECT FINANCIAL REPORT"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Project Information
        row = 3
        ws[f'A{row}'] = "Project:"
        ws[f'B{row}'] = project.get('title', 'N/A')
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Project ID:"
        ws[f'B{row}'] = project.get('project_no', 'N/A')
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Principal Investigator:"
        ws[f'B{row}'] = project.get('pi_name', 'N/A')
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Technical Group:"
        ws[f'B{row}'] = project.get('technical_group_name', 'N/A')
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Duration:"
        ws[f'B{row}'] = f"{ExcelReportGenerator.format_date(project.get('start_date'))} to {ExcelReportGenerator.format_date(project.get('end_date'))}"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%d-%b-%Y %I:%M %p")
        ws[f'A{row}'].font = Font(bold=True)
        
        # Financial Summary
        row += 3
        ws[f'A{row}'] = "FINANCIAL SUMMARY"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        # Budget Overview
        row += 2
        ws[f'A{row}'] = "Budget Overview"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Approved Budget:"
        ws[f'B{row}'] = ExcelReportGenerator.format_currency(summary.get('total_budget', 0))
        ws[f'B{row}'].number_format = '₹#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "Committed:"
        ws[f'B{row}'] = ExcelReportGenerator.format_currency(summary.get('total_committed', 0))
        ws[f'B{row}'].number_format = '₹#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "Balance:"
        ws[f'B{row}'] = ExcelReportGenerator.format_currency(summary.get('budget_balance', 0))
        ws[f'B{row}'].number_format = '₹#,##0.00'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Utilization (%):"
        utilization_value = summary.get('budget_utilization', 0) / 100
        ws[f'B{row}'] = utilization_value
        ws[f'B{row}'].number_format = '0.00%'
        
        # Funds Overview
        row += 2
        ws[f'A{row}'] = "Funds Overview"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Funds Received:"
        ws[f'B{row}'] = ExcelReportGenerator.format_currency(summary.get('total_funds', 0))
        ws[f'B{row}'].number_format = '₹#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "Expenditure:"
        ws[f'B{row}'] = ExcelReportGenerator.format_currency(summary.get('total_spent', 0))
        ws[f'B{row}'].number_format = '₹#,##0.00'
        
        row += 1
        ws[f'A{row}'] = "Balance:"
        ws[f'B{row}'] = ExcelReportGenerator.format_currency(summary.get('funds_balance', 0))
        ws[f'B{row}'].number_format = '₹#,##0.00'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Utilization (%):"
        funds_utilization_value = summary.get('funds_utilization', 0) / 100
        ws[f'B{row}'] = funds_utilization_value
        ws[f'B{row}'].number_format = '0.00%'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    @staticmethod
    def _create_budget_sheet(wb: Workbook, data: Dict[str, Any]):
        """Create Budget Allocation sheet"""
        
        ws = wb.create_sheet("Budget Allocation")
        budget_data = data.get('budget_allocation', [])
        
        # Title
        ws['A1'] = "BUDGET ALLOCATION BY CATEGORY"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ['Category', 'Approved Budget', 'Committed', 'Balance', 'Utilization (%)']
        header_row = 3
        
        # Style for headers
        header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = header_row + 1
        total_approved = 0
        total_committed = 0
        total_balance = 0
        
        for item in budget_data:
            approved = ExcelReportGenerator.format_currency(item.get('approved_budget', 0))
            committed = ExcelReportGenerator.format_currency(item.get('committed_amount', 0))
            balance = ExcelReportGenerator.format_currency(item.get('balance', 0))
            utilization = item.get('utilization_percentage', 0) / 100
            
            total_approved += approved
            total_committed += committed
            total_balance += balance
            
            ws[f'A{row}'] = str(item.get('category', 'N/A')).title()
            ws[f'B{row}'] = approved
            ws[f'C{row}'] = committed
            ws[f'D{row}'] = balance
            ws[f'E{row}'] = utilization
            
            # Formatting
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
            
            ws[f'B{row}'].number_format = '₹#,##0.00'
            ws[f'C{row}'].number_format = '₹#,##0.00'
            ws[f'D{row}'].number_format = '₹#,##0.00'
            ws[f'E{row}'].number_format = '0.00%'
            
            row += 1
        
        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = total_approved
        ws[f'C{row}'] = total_committed
        ws[f'D{row}'] = total_balance
        ws[f'E{row}'] = (total_committed / total_approved) if total_approved > 0 else 0
        
        # Style total row
        total_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.border = thin_border
        
        ws[f'B{row}'].number_format = '₹#,##0.00'
        ws[f'C{row}'].number_format = '₹#,##0.00'
        ws[f'D{row}'].number_format = '₹#,##0.00'
        ws[f'E{row}'].number_format = '0.00%'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 18
    
    @staticmethod
    def _create_funds_sheet(wb: Workbook, data: Dict[str, Any]):
        """Create Funds & Expenditure sheet"""
        
        ws = wb.create_sheet("Funds & Expenditure")
        funds_data = data.get('funds_expenditure', [])
        
        # Title
        ws['A1'] = "FUNDS RECEIVED & EXPENDITURE BY CATEGORY"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ['Category', 'Funds Received', 'Spent', 'Balance', 'Utilization (%)']
        header_row = 3
        
        # Style for headers
        header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = header_row + 1
        total_received = 0
        total_spent = 0
        total_balance = 0
        
        for item in funds_data:
            received = ExcelReportGenerator.format_currency(item.get('funds_received', 0))
            spent = ExcelReportGenerator.format_currency(item.get('spent', 0))
            balance = ExcelReportGenerator.format_currency(item.get('balance', 0))
            utilization = (spent / received) if received > 0 else 0
            
            total_received += received
            total_spent += spent
            total_balance += balance
            
            ws[f'A{row}'] = str(item.get('category', 'N/A')).title()
            ws[f'B{row}'] = received
            ws[f'C{row}'] = spent
            ws[f'D{row}'] = balance
            ws[f'E{row}'] = utilization
            
            # Formatting
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
            
            ws[f'B{row}'].number_format = '₹#,##0.00'
            ws[f'C{row}'].number_format = '₹#,##0.00'
            ws[f'D{row}'].number_format = '₹#,##0.00'
            ws[f'E{row}'].number_format = '0.00%'
            
            row += 1
        
        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = total_received
        ws[f'C{row}'] = total_spent
        ws[f'D{row}'] = total_balance
        ws[f'E{row}'] = (total_spent / total_received) if total_received > 0 else 0
        
        # Style total row
        total_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.border = thin_border
        
        ws[f'B{row}'].number_format = '₹#,##0.00'
        ws[f'C{row}'].number_format = '₹#,##0.00'
        ws[f'D{row}'].number_format = '₹#,##0.00'
        ws[f'E{row}'].number_format = '0.00%'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 18
    
    @staticmethod
    def _create_transaction_sheets(wb: Workbook, data: Dict[str, Any]):
        """Create detailed transaction sheets for each category"""
        
        categories = data.get('categories', {})
        # Implementation remains the same...
        pass
    
    @staticmethod
    def generate_projects_summary_excel(projects_data: list) -> str:
        """Generate Excel report for all projects summary"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Projects Summary"
        
        # Title
        ws['A1'] = "ALL PROJECTS SUMMARY"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:J1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Subtitle with date
        ws['A2'] = f"Generated on: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:J2')
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = [
            'Project No', 'Title', 'Technical Group', 'Funding Agency',
            'Approved Budget', 'Funds Received', 'Expenditure',
            'Budget Balance', 'Funds Balance', 'Utilization'
        ]
        header_row = 4
        
        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Data rows
        row = header_row + 1
        total_budget = 0
        total_funds = 0
        total_expenditure = 0
        total_budget_balance = 0
        total_funds_balance = 0
        
        for project in projects_data:
            approved_budget = ExcelReportGenerator.format_currency(project.get('approved_budget', 0))
            funds_received = ExcelReportGenerator.format_currency(project.get('funds_received', 0))
            expenditure = ExcelReportGenerator.format_currency(project.get('expenditure', 0))
            budget_balance = ExcelReportGenerator.format_currency(project.get('budget_balance', 0))
            funds_balance = ExcelReportGenerator.format_currency(project.get('funds_balance', 0))
            utilization = project.get('utilization', 0) / 100  # Divide by 100 for percentage format
            
            total_budget += approved_budget
            total_funds += funds_received
            total_expenditure += expenditure
            total_budget_balance += budget_balance
            total_funds_balance += funds_balance
            
            ws[f'A{row}'] = project.get('project_no', 'N/A')
            ws[f'B{row}'] = project.get('title', 'N/A')
            ws[f'C{row}'] = project.get('technical_group', 'N/A')
            ws[f'D{row}'] = project.get('funding_agency', 'N/A')
            ws[f'E{row}'] = approved_budget
            ws[f'F{row}'] = funds_received
            ws[f'G{row}'] = expenditure
            ws[f'H{row}'] = budget_balance
            ws[f'I{row}'] = funds_balance
            ws[f'J{row}'] = utilization
            
            # Formatting
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
            
            # Currency formatting
            ws[f'E{row}'].number_format = '₹#,##0'
            ws[f'F{row}'].number_format = '₹#,##0'
            ws[f'G{row}'].number_format = '₹#,##0'
            ws[f'H{row}'].number_format = '₹#,##0'
            ws[f'I{row}'].number_format = '₹#,##0'
            ws[f'J{row}'].number_format = '0.0%'
            
            # Color coding for balances
            if budget_balance > 0:
                ws[f'H{row}'].font = Font(color="10B981")  # Green
            if funds_balance > 0:
                ws[f'I{row}'].font = Font(color="10B981")  # Green
            
            row += 1
        
        # Total row
        ws[f'A{row}'] = "TOTAL"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        
        ws[f'E{row}'] = total_budget
        ws[f'F{row}'] = total_funds
        ws[f'G{row}'] = total_expenditure
        ws[f'H{row}'] = total_budget_balance
        ws[f'I{row}'] = total_funds_balance
        
        avg_utilization = (total_expenditure / total_funds) if total_funds > 0 else 0
        ws[f'J{row}'] = avg_utilization
        
        # Style total row
        total_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = total_fill
            cell.border = thin_border
        
        ws[f'E{row}'].number_format = '₹#,##0'
        ws[f'F{row}'].number_format = '₹#,##0'
        ws[f'G{row}'].number_format = '₹#,##0'
        ws[f'H{row}'].number_format = '₹#,##0'
        ws[f'I{row}'].number_format = '₹#,##0'
        ws[f'J{row}'].number_format = '0.0%'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 16
        ws.column_dimensions['I'].width = 16
        ws.column_dimensions['J'].width = 12
        
        # Freeze header row
        ws.freeze_panes = 'A5'
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        temp_file.close()
        
        wb.save(temp_path)
        return temp_path